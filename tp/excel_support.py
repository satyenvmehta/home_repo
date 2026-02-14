from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Any, Tuple
import re

# ============================================================
# Enums (no magic strings)
# ============================================================

class FillColor(Enum):
    YELLOW = "FFFF00"
    GREEN  = "CCFFCC"
    RED    = "FFC7CE"
    BLUE   = "ADD8E6"


class ConditionOp(Enum):
    GT = ">"
    LT = "<"
    GTE = ">="
    LTE = "<="
    EQ = "=="
    BETWEEN = "between"


@dataclass(frozen=True)
class Condition:
    op: ConditionOp
    value: Any
    color: FillColor


# ============================================================
# Shared helpers
# ============================================================

def cell_to_rc(cell: str) -> Tuple[int, int]:
    match = re.match(r"([A-Z]+)(\d+)", cell.upper())
    if not match:
        raise ValueError(f"Invalid cell reference: {cell}")

    col_letters, row = match.groups()
    col = 0
    for c in col_letters:
        col = col * 26 + (ord(c) - ord("A") + 1)

    return int(row) - 1, col - 1


def col_letter(col: int) -> str:
    """1-based column index → Excel column letter"""
    result = ""
    while col:
        col, rem = divmod(col - 1, 26)
        result = chr(65 + rem) + result
    return result


# ============================================================
# ExcelCreator (CREATE-TIME, xlsxwriter)
# ============================================================

import pandas as pd

@dataclass
class ExcelCreator:
    writer: pd.ExcelWriter

    def __post_init__(self):
        self.workbook = self.writer.book
        self.sheets = {}
        self.formats = {}

    def add_sheet(self, name: str, df: pd.DataFrame):
        df.to_excel(self.writer, sheet_name=name, index=False)
        self.sheets[name] = self.writer.sheets[name]

    def _fmt(self, color: FillColor):
        if color not in self.formats:
            self.formats[color] = self.workbook.add_format(
                {"bg_color": f"#{color.value}"}
            )
        return self.formats[color]

    # -------- formatting --------

    def fill_cell(self, sheet, cell, color):
        ws = self.sheets[sheet]
        r, c = cell_to_rc(cell)
        ws.write(r, c, "", self._fmt(color))

    def fill_range(self, sheet, cell_range, color):
        ws = self.sheets[sheet]
        ws.conditional_format(
            cell_range,
            {"type": "no_errors", "format": self._fmt(color)}
        )

    def fill_row_for_df(self, sheet, row: int, df, color):
        last_col = df.shape[1]
        rng = f"A{row}:{col_letter(last_col)}{row}"
        self.fill_range(sheet, rng, color)

    def fill_row(self, sheet, row, color):
        ws = self.sheets[sheet]
        ws.set_row(row - 1, None, self._fmt(color))

    def fill_col(self):
        raise NotImplementedError

    def conditional_format(self, sheet, cell_range, condition: Condition):
        ws = self.sheets[sheet]
        fmt = self._fmt(condition.color)

        if condition.op == ConditionOp.BETWEEN:
            low, high = condition.value
            ws.conditional_format(cell_range, {
                "type": "cell",
                "criteria": "between",
                "minimum": low,
                "maximum": high,
                "format": fmt
            })
        else:
            ws.conditional_format(cell_range, {
                "type": "cell",
                "criteria": condition.op.value,
                "value": condition.value,
                "format": fmt
            })


# ============================================================
# ExcelUpdater (UPDATE-TIME, openpyxl)
# ============================================================

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

@dataclass
class ExcelUpdater:
    path: Path

    def __post_init__(self):
        self.wb = load_workbook(self.path)

    def _fill(self, color: FillColor):
        return PatternFill(
            start_color=color.value,
            end_color=color.value,
            fill_type="solid"
        )

    def fill_cell(self, sheet, cell, color):
        self.wb[sheet][cell].fill = self._fill(color)

    def fill_range(self, sheet, cell_range, color):
        ws = self.wb[sheet]
        fill = self._fill(color)
        for row in ws[cell_range]:
            for cell in row:
                cell.fill = fill

    def fill_row(self, sheet, row, color):
        ws = self.wb[sheet]
        fill = self._fill(color)
        for c in range(1, ws.max_column + 1):
            ws.cell(row=row, column=c).fill = fill

    def conditional_format(self, sheet, cell_range, condition: Condition):
        ws = self.wb[sheet]
        fill = self._fill(condition.color)

        if condition.op == ConditionOp.BETWEEN:
            low, high = condition.value
            rule = CellIsRule(
                operator="between",
                formula=[str(low), str(high)],
                fill=fill
            )
        else:
            rule = CellIsRule(
                operator={
                    ConditionOp.GT: "greaterThan",
                    ConditionOp.LT: "lessThan",
                    ConditionOp.GTE: "greaterThanOrEqual",
                    ConditionOp.LTE: "lessThanOrEqual",
                    ConditionOp.EQ: "equal",
                }[condition.op],
                formula=[str(condition.value)],
                fill=fill
            )

        ws.conditional_formatting.add(cell_range, rule)

    def save(self):
        self.wb.save(self.path)


import pandas as pd
from pathlib import Path
# from excel_support import *

if __name__ == "__main__":
    import sys
    print("Usage: from excel_support import *")

    df_summary = pd.DataFrame({
        "Metric": ["Revenue", "Cost", "Profit"],
        "Value": [1200, 800, 400]
    })


    def apply_formatter(excel, df):
        excel.fill_row_for_df("Summary", 1, df, FillColor.YELLOW)
        excel.conditional_format(
            "Summary",
            "B2:B10",
            Condition(ConditionOp.GT, 1000, FillColor.RED)
        )

    path = Path("C:/tmp/report.xlsx")

    with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
        excel = ExcelCreator(path=path, writer=writer)
        excel.add_sheet("Summary", df_summary)
        apply_formatter(excel, df_summary)
    # file is created here
