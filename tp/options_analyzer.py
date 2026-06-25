"""
Stock & Options Batch Analyzer
================================
Feed it a list of tickers, get back a formatted Excel spreadsheet with
live prices, trend signals, and suggested stock + option trades sized
for a $500 per-trade budget.

SETUP (one time):
    pip install yfinance pandas openpyxl

USAGE:
    python options_analyzer.py AAPL MSFT PFE MU NVDA
    (or)
    python options_analyzer.py tickers.txt     # one ticker per line

OUTPUT:
    trade_ideas_YYYY-MM-DD_HHMM.xlsx  (in the same folder)

DISCLAIMER: For educational purposes only. Not financial advice.
Always verify prices and options data in Fidelity before trading.
"""

import sys
import datetime as dt
from pathlib import Path

import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BUDGET = 500  # dollars per trade
STOCK_TARGET_PCT = 5
STOCK_STOP_PCT = 3
OPTION_TARGET_PCT = 25  # gain on premium
OPTION_STOP_PCT = 20    # loss on premium


def get_tickers():
    args = ["T", "PFE"] #sys.argv[1:]
    if not args:
        print("Usage: python options_analyzer.py TICKER1 TICKER2 ...")
        print("   or: python options_analyzer.py tickers.txt")
        sys.exit(1)
    if len(args) == 1 and Path(args[0]).exists():
        with open(args[0]) as f:
            return [line.strip().upper() for line in f if line.strip()]
    return [t.strip().upper() for t in args]


def compute_rsi(close, period=14):
    delta = close.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.rolling(period).mean()
    avg_loss = loss.rolling(period).mean()
    rs = avg_gain / avg_loss
    rsi = 100 - (100 / (1 + rs))
    return rsi.iloc[-1]


def pick_expiration(ticker_obj, min_days=4, max_days=14):
    """Pick the nearest real expiration date within the day window."""
    today = dt.date.today()
    for exp_str in ticker_obj.options:
        exp_date = dt.datetime.strptime(exp_str, "%Y-%m-%d").date()
        days_out = (exp_date - today).days
        if min_days <= days_out <= max_days:
            return exp_str
    # fallback: first available future expiration
    for exp_str in ticker_obj.options:
        exp_date = dt.datetime.strptime(exp_str, "%Y-%m-%d").date()
        if (exp_date - today).days >= min_days:
            return exp_str
    return ticker_obj.options[0] if ticker_obj.options else None


def analyze_ticker(symbol):
    row = {"Ticker": symbol}
    try:
        t = yf.Ticker(symbol)
        hist = t.history(period="6mo")
        if hist.empty:
            row["Error"] = "No price data found"
            return row

        price = hist["Close"].iloc[-1]
        prev_close = hist["Close"].iloc[-2]
        change_pct = (price - prev_close) / prev_close * 100

        sma20 = hist["Close"].rolling(20).mean().iloc[-1]
        sma50 = hist["Close"].rolling(50).mean().iloc[-1] if len(hist) >= 50 else sma20
        rsi = compute_rsi(hist["Close"])

        high_52w = hist["Close"].max()
        low_52w = hist["Close"].min()
        pct_from_high = (price - high_52w) / high_52w * 100

        # --- Trend logic ---
        if price > sma20 > sma50 and rsi < 70:
            trend = "Bullish"
        elif price < sma20 < sma50 and rsi > 30:
            trend = "Bearish"
        else:
            trend = "Neutral"

        # --- Risk level (based on volatility) ---
        daily_returns = hist["Close"].pct_change().dropna()
        volatility = daily_returns.std() * (252 ** 0.5) * 100  # annualized %
        if volatility < 30:
            risk = "Low"
        elif volatility < 55:
            risk = "Medium"
        else:
            risk = "High"

        # --- Stock trade suggestion ---
        if trend == "Bullish" and rsi < 70:
            stock_action = "BUY"
            entry = round(price, 2)
            target = round(price * (1 + STOCK_TARGET_PCT / 100), 2)
            stop = round(price * (1 - STOCK_STOP_PCT / 100), 2)
        elif trend == "Bearish" and rsi > 30:
            stock_action = "SELL/AVOID"
            entry = target = stop = None
        elif rsi < 30:
            stock_action = "WATCH (oversold bounce?)"
            entry = target = stop = None
        elif rsi > 70:
            stock_action = "AVOID (overbought)"
            entry = target = stop = None
        else:
            stock_action = "NEUTRAL"
            entry = target = stop = None

        # --- Option trade suggestion (real chain data) ---
        opt_type = strike = exp = premium = contracts = total_cost = None
        opt_target = opt_stop = None
        try:
            exp = pick_expiration(t)
            if exp:
                chain = t.option_chain(exp)
                opt_type = "CALL" if trend == "Bullish" else ("PUT" if trend == "Bearish" else "CALL")
                table = chain.calls if opt_type == "CALL" else chain.puts

                # pick strike nearest to current price (ATM)
                table = table.copy()
                table["diff"] = (table["strike"] - price).abs()
                best = table.sort_values("diff").iloc[0]
                strike = best["strike"]
                bid, ask = best["bid"], best["ask"]
                premium = round((bid + ask) / 2, 2) if (bid and ask) else round(best["lastPrice"], 2)

                if premium and premium > 0:
                    contracts = max(1, int(BUDGET // (premium * 100)))
                    total_cost = round(contracts * premium * 100, 2)
                    opt_target = round(premium * (1 + OPTION_TARGET_PCT / 100), 2)
                    opt_stop = round(premium * (1 - OPTION_STOP_PCT / 100), 2)
        except Exception:
            pass  # options data not available for this ticker

        row.update({
            "Price": round(price, 2),
            "Chg %": round(change_pct, 2),
            "52W Low": round(low_52w, 2),
            "52W High": round(high_52w, 2),
            "% From 52W High": round(pct_from_high, 1),
            "RSI(14)": round(rsi, 1) if pd.notna(rsi) else None,
            "Trend": trend,
            "Risk": risk,
            "Stock Action": stock_action,
            "Stock Entry": entry,
            "Stock Target": target,
            "Stock Stop": stop,
            "Option Type": opt_type,
            "Option Strike": strike,
            "Option Exp": exp,
            "Option Premium": premium,
            "Contracts ($500 budget)": contracts,
            "Option Cost": total_cost,
            "Option Target Premium": opt_target,
            "Option Stop Premium": opt_stop,
        })
    except Exception as e:
        row["Error"] = str(e)
    return row


def build_spreadsheet(rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Ideas"

    headers = [
        "Ticker", "Price", "Chg %", "52W Low", "52W High", "% From 52W High",
        "RSI(14)", "Trend", "Risk", "Stock Action", "Stock Entry", "Stock Target",
        "Stock Stop", "Option Type", "Option Strike", "Option Exp",
        "Option Premium", "Contracts ($500 budget)", "Option Cost",
        "Option Target Premium", "Option Stop Premium", "Error",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", start_color="1F2937")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    trend_colors = {"Bullish": "DCFCE7", "Bearish": "FEE2E2", "Neutral": "F1F5F9"}
    risk_colors = {"Low": "DCFCE7", "Medium": "FEF3C7", "High": "FEE2E2"}
    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_data in rows:
        ws.append([row_data.get(h, "") for h in headers])
        r = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        trend = row_data.get("Trend")
        if trend in trend_colors:
            ws.cell(row=r, column=headers.index("Trend") + 1).fill = PatternFill("solid", start_color=trend_colors[trend])

        risk = row_data.get("Risk")
        if risk in risk_colors:
            ws.cell(row=r, column=headers.index("Risk") + 1).fill = PatternFill("solid", start_color=risk_colors[risk])

        action = row_data.get("Stock Action", "")
        if "BUY" in action:
            ws.cell(row=r, column=headers.index("Stock Action") + 1).fill = PatternFill("solid", start_color="DCFCE7")
        elif "AVOID" in action or "SELL" in action:
            ws.cell(row=r, column=headers.index("Stock Action") + 1).fill = PatternFill("solid", start_color="FEE2E2")

    # column widths
    widths = [8, 9, 7, 8, 8, 14, 8, 9, 8, 16, 11, 12, 10, 10, 12, 12, 12, 18, 11, 16, 16, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    # Notes sheet
    notes = wb.create_sheet("Notes")
    notes_text = [
        ["Stock & Options Batch Analyzer - Notes"],
        [""],
        [f"Generated: {dt.datetime.now().strftime('%Y-%m-%d %H:%M')}"],
        [f"Budget per trade: ${BUDGET}"],
        [f"Stock target / stop: +{STOCK_TARGET_PCT}% / -{STOCK_STOP_PCT}%"],
        [f"Option target / stop (on premium): +{OPTION_TARGET_PCT}% / -{OPTION_STOP_PCT}%"],
        [""],
        ["Trend logic:"],
        ["  Bullish = price > SMA20 > SMA50 and RSI < 70"],
        ["  Bearish = price < SMA20 < SMA50 and RSI > 30"],
        ["  Neutral = everything else"],
        [""],
        ["Risk = annualized volatility: Low <30%, Medium 30-55%, High >55%"],
        [""],
        ["IMPORTANT: This is a heuristic screen, not financial advice."],
        ["Always verify current prices and option quotes in Fidelity before trading."],
        ["Option premiums are mid (bid+ask)/2 at time of run - they move fast."],
    ]
    for line in notes_text:
        notes.append(line)
    notes["A1"].font = Font(bold=True, size=13, name="Arial")
    for r in range(2, len(notes_text) + 1):
        notes.cell(row=r, column=1).font = Font(name="Arial", size=10)
    notes.column_dimensions["A"].width = 70

    wb.save(out_path)


def main():
    tickers = get_tickers()
    print(f"Analyzing {len(tickers)} ticker(s): {', '.join(tickers)}")
    rows = []
    for sym in tickers:
        print(f"  -> {sym} ...")
        rows.append(analyze_ticker(sym))

    timestamp = dt.datetime.now().strftime("%Y-%m-%d_%H%M")
    out_path = f"trade_ideas_{timestamp}.xlsx"
    build_spreadsheet(rows, out_path)
    print(f"\nDone! Saved to: {out_path}")


if __name__ == "__main__":
    main()
