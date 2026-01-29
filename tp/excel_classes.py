# ============================================================
# excel_framework.py
# ============================================================

from abc import ABC, abstractmethod
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Any, Callable, Tuple
import re

# ============================================================
# Enums
# ============================================================

class FillColor(Enum):
    YELLOW = "FFFF00"
    GREEN  = "CCFFCC"
    RED    = "FFC7CE"
    BLUE   = "ADD8E6"


class ConditionOp(Enum):
    GT = ">"
    LT = "<"
    GTE = ">="
    LTE = "<="
    EQ = "=="
    BETWEEN = "between"


# ============================================================
# Condition (intent-only)
# ============================================================

@dataclass(frozen=True)
class Condition:
    op: ConditionOp
    value: Any
    color: FillColor


# ============================================================
# Base abstraction
# ============================================================

@dataclass
class ExcelFileBase(ABC):
    path: Path

    # ---------- utilities ----------
    @staticmethod
    def cell_to_rc(cell: str) -> Tuple[int, int]:
        match = re.match(r"([A-Z]+)(\d+)", cell.upper())
        if not match:
            raise ValueError(f"Invalid cell reference: {cell}")

        col_letters, row = match.groups()
        col = 0
        for c in col_letters:
            col = col * 26 + (ord(c) - ord("A") + 1)

        return int(row) - 1, col - 1

    # ---------- formatter hook ----------
    def apply_formatter(self, formatter: Callable[["ExcelFileBase"], None]) -> None:
        formatter(self)

    # ---------- abstract API ----------
    @abstractmethod
    def fill_cell(self, sheet: str, cell: str, color: FillColor): ...
    @abstractmethod
    def fill_row(self, sheet: str, row: int, color: FillColor): ...
    @abstractmethod
    def fill_range(self, sheet: str, cell_range: str, color: FillColor): ...
    @abstractmethod
    def conditional_format(self, sheet: str, cell_range: str, condition: Condition): ...
    @abstractmethod
    def save(self): ...


# ============================================================
# ExcelCreator (xlsxwriter, create-time)
# ============================================================

import pandas as pd

@dataclass
class ExcelCreator(ExcelFileBase):
    writer: pd.ExcelWriter

    def __post_init__(self):
        self.workbook = self.writer.book
        self._sheets = {}
        self._formats = {}

    def add_sheet(self, name: str, df: pd.DataFrame):
        df.to_excel(self.writer, sheet_name=name, index=False)
        self._sheets[name] = self.writer.sheets[name]

    # ---------- helpers ----------
    def _get_format(self, color: FillColor):
        if color not in self._formats:
            self._formats[color] = self.workbook.add_format({
                "bg_color": f"#{color.value}"
            })
        return self._formats[color]

    # ---------- fills ----------
    def fill_cell(self, sheet, cell, color):
        ws = self._sheets[sheet]
        r, c = self.cell_to_rc(cell)
        ws.write(r, c, "", self._get_format(color))

    def fill_row(self, sheet, row, color):
        # STRUCTURAL: applies to entire row
        ws = self._sheets[sheet]
        ws.set_row(row - 1, None, self._get_format(color))

    def fill_range(self, sheet, cell_range, color):
        ws = self._sheets[sheet]
        ws.conditional_format(cell_range, {
            "type": "no_errors",
            "format": self._get_format(color)
        })

    # ---------- conditional formatting ----------
    def conditional_format(self, sheet, cell_range, condition):
        ws = self._sheets[sheet]
        fmt = self._get_format(condition.color)

        if condition.op == ConditionOp.BETWEEN:
            low, high = condition.value
            ws.conditional_format(cell_range, {
                "type": "cell",
                "criteria": "between",
                "minimum": low,
                "maximum": high,
                "format": fmt
            })
        else:
            ws.conditional_format(cell_range, {
                "type": "cell",
                "criteria": condition.op.value,
                "value": condition.value,
                "format": fmt
            })

    def save(self):
        self.writer.close()


# ============================================================
# ExcelUpdater (openpyxl, post-write)
# ============================================================

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

@dataclass
class ExcelUpdater(ExcelFileBase):

    def __post_init__(self):
        self.wb = load_workbook(self.path)

    def _fill(self, color: FillColor):
        return PatternFill(
            start_color=color.value,
            end_color=color.value,
            fill_type="solid"
        )

    # ---------- fills ----------
    def fill_cell(self, sheet, cell, color):
        self.wb[sheet][cell].fill = self._fill(color)

    def fill_row(self, sheet, row, color):
        ws = self.wb[sheet]
        fill = self._fill(color)
        for c in range(1, ws.max_column + 1):
            ws.cell(row=row, column=c).fill = fill

    def fill_range(self, sheet, cell_range, color):
        ws = self.wb[sheet]
        fill = self._fill(color)
        for row in ws[cell_range]:
            for cell in row:
                cell.fill = fill

    # ---------- conditional formatting ----------
    _OP_MAP = {
        ConditionOp.GT: "greaterThan",
        ConditionOp.LT: "lessThan",
        ConditionOp.GTE: "greaterThanOrEqual",
        ConditionOp.LTE: "lessThanOrEqual",
        ConditionOp.EQ: "equal",
    }

    def conditional_format(self, sheet, cell_range, condition):
        ws = self.wb[sheet]
        fill = self._fill(condition.color)

        if condition.op == ConditionOp.BETWEEN:
            low, high = condition.value
            rule = CellIsRule(
                operator="between",
                formula=[str(low), str(high)],
                fill=fill
            )
        else:
            rule = CellIsRule(
                operator=self._OP_MAP[condition.op],
                formula=[str(condition.value)],
                fill=fill
            )

        ws.conditional_formatting.add(cell_range, rule)

    def save(self):
        self.wb.save(self.path)

def create_excel(full_path: Path, df: pd.DataFrame, sheet_name: str = "Sheet1"):
    with pd.ExcelWriter(full_path, engine="xlsxwriter") as writer:
        excel = ExcelCreator(path=full_path, writer=writer)
        excel.add_sheet("Summary", df_summary)
        excel.apply_formatter(formatter)

        df.to_excel(writer, sheet_name=sheet_name, index=False)
        excel = ExcelCreator(path=full_path, writer=writer)
        return excel

def create_excel_file(full_path: Path, mode: str = "create") -> ExcelFileBase:
    if mode == "create":
        with pd.ExcelWriter(full_path, engine="xlsxwriter") as writer:
            excel = ExcelCreator(path=full_path, writer=writer)
            return excel
    elif mode == "update":
        return ExcelUpdater(path=full_path)
    else:
        raise ValueError(f"Unknown mode: {mode}")

