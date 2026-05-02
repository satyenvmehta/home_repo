# ============================================================
# excel_updater.py
# ============================================================

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Border, Side

from excel_base import ExcelFileBase, FillColor, Condition, ConditionOp


@dataclass
class ExcelUpdater(ExcelFileBase):
    """
    Update-time Excel handler backed by openpyxl.
    Works on existing .xlsx files.
    """

    def __post_init__(self):
        self.workbook = load_workbook(self.path)
        # populate sheets dict
        self.sheets = {name: self.workbook[name] for name in self.workbook.sheetnames}

        # reusable border object
        thin = Side(style="thin")
        self._thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ----------------------------
    # Helpers
    # ----------------------------
    def _fill(self, color: FillColor) -> PatternFill:
        return PatternFill(start_color=color.value, end_color=color.value, fill_type="solid")

    # ----------------------------
    # Base API implementations
    # ----------------------------
    def fill_cell(self, sheet: str, cell: str, color: FillColor) -> None:
        ws = self.get_sheet(sheet)
        ws[cell].fill = self._fill(color)

    def fill_range(self, sheet: str, cell_range: str, color: FillColor) -> None:
        ws = self.get_sheet(sheet)
        fill = self._fill(color)
        for row in ws[cell_range]:
            for cell in row:
                cell.fill = fill

    def fill_row(self, sheet: str, row: int, color: FillColor) -> None:
        ws = self.get_sheet(sheet)
        fill = self._fill(color)
        for c in range(1, ws.max_column + 1):
            ws.cell(row=row, column=c).fill = fill

    def fill_column(self, sheet: str, col: int, color: FillColor) -> None:
        ws = self.get_sheet(sheet)
        fill = self._fill(color)
        for r in range(1, ws.max_row + 1):
            ws.cell(row=r, column=col).fill = fill

    def conditional_format(self, sheet: str, cell_range: str, condition: Condition) -> None:
        ws = self.get_sheet(sheet)
        fill = self._fill(condition.color)

        op_map = {
            ConditionOp.GT: "greaterThan",
            ConditionOp.LT: "lessThan",
            ConditionOp.GTE: "greaterThanOrEqual",
            ConditionOp.LTE: "lessThanOrEqual",
            ConditionOp.EQ: "equal",
        }

        if condition.op == ConditionOp.BETWEEN:
            low, high = condition.value
            rule = CellIsRule(operator="between", formula=[str(low), str(high)], fill=fill)
        elif condition.op == ConditionOp.CONTAINS:
            ws.conditional_format(cell_range, {
                "type": "text",
                "criteria": "containing",
                "value": condition.value,
                "format": fmt
            })
        else:
            rule = CellIsRule(operator=op_map[condition.op], formula=[str(condition.value)], fill=fill)

        ws.conditional_formatting.add(cell_range, rule)

    def border_used_range(self, sheet: str) -> None:
        ws = self.get_sheet(sheet)
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row <= 0 or max_col <= 0:
            return

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).border = self._thin_border

    def save(self) -> None:
        self.workbook.save(self.path)
