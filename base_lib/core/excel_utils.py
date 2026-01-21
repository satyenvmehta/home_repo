
from base_lib.core.common_include import *
from base_lib.core.base_classes import BaseObject
from base_lib.core.base_container_classes import  BaseList, BaseDict

from openpyxl import Workbook
from openpyxl.styles import PatternFill

def fromColName2ColIndex(df, colName):
    char_col_index = None
    if isinstance(df, DataFrame):
        col_index = df.columns.get_loc(colName)  # Get the index of the 'Amount' column
        char_col_index = chr(col_index + 65)
    return char_col_index

@dataclass
class BaseExcel(BaseObject):
    sheetNames: Any = None

    def __post_init__(self):
        wb = Workbook()
        self.mainsheet = wb.active
        self.setBase(wb)
        self.getBase().active
        self.sheetDict = BaseDict()
        return

    def getWorkbook(self):
        return self.getBase()

    def getMainSheet(self):
        return self.mainsheet

    def addToSheetDict(self, name, sheet):
        self.sheetDict.append(name, sheet)
        return

    def setMainSheet(self):
        self.mainsheet = self.getWorkbook().active
        self.mainsheet.title = 'Main'
        self.addToSheetDict(self.getMainSheet(), self.mainsheet.title)
        return

    def addSheet(self, name):
        self.getWorkbook().create_sheet(title='Sheet2')
    # def setSheetNames(self):
    #     wb = self.getBase()
    #     if isinstance(wb, Workbook):
    #         names = wb.sheetnames
    #         self.sheetNames = BaseList()
    #         self.sheetNames.setBase(names)
    #     return

    def getSheetNames(self):
        if self.sheetNames:
            return self.sheetNames
        self.setSheetNames()
        return self.sheetNames

    def setTitle(self, title="main"):
        pass

# @dataclass
# class BaseExcelExisting(BaseExcel):
#     fileName : BaseObject
#
#     def __post_init__(self):
#         wb = load_workbook(self.fileName)
#         lst = self.getSheetNames()
#         lst.print()
#         return


def testExceBasic(be):
    sheet = be.getMainSheet()
    # Add some data to the sheet
    sheet['A1'] = 'Click here to go to Sheet2'
    sheet['B1'] = 'Data in another tab'
    workbook = be.getWorkbook()

    # Create a new sheet
    sheet2 = workbook.create_sheet(title='Sheet2')

    # Add some data to the second sheet
    sheet2['A1'] = 'Data in Sheet2'

    # Create a hyperlink from A1 to cell A1 in Sheet2
    sheet['A1'].hyperlink = f"#'Sheet2'!A1"

    # Save the workbook
    workbook.save('example.xlsx')
    print("Excel file created successfully, " + "example.xlsx")
    return

if  __name__ == "__main__":
    # Create a new Excel workbook
    be = BaseExcel()
    testExceBasic(be)  # Need to be tested
